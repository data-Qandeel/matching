import os
import logging
import re
import unicodedata
from openpyxl import load_workbook
import pandas as pd
import torch
from typing import Optional, Dict, List, Tuple, Union, Any
from pathlib import Path


logger = logging.getLogger(__name__)
LOGGING_LEVEL = logging.INFO


def setup_logging(
    log_file: Optional[Union[str, Path]] = None,
    log_level: int = LOGGING_LEVEL,
    log_format: str = "[%(asctime)s] %(levelname)s - %(name)s - %(message)s",
    date_format: str = "%Y-%m-%d %H:%M:%S",
) -> None:
    """ Configure logging to console and optionally to a file. """
    root_logger = logging.getLogger()
    root_logger.setLevel(log_level)

    # Clear existing handlers
    for handler in root_logger.handlers[:]:
        root_logger.removeHandler(handler)
        handler.close()

    # Create and add console handler
    console_handler = logging.StreamHandler()
    console_handler.setFormatter(logging.Formatter(log_format, datefmt=date_format))
    root_logger.addHandler(console_handler)

    # Add file handler if specified
    if log_file:
        try:
            log_path = Path(log_file)
            log_path.parent.mkdir(parents=True, exist_ok=True)

            file_handler = logging.FileHandler(log_path, mode="a", encoding="utf-8")
            file_handler.setFormatter(logging.Formatter(log_format, datefmt=date_format))
            root_logger.addHandler(file_handler)

            root_logger.info(f"Logging configured. Outputting to console and file: {log_path}")
        except Exception as e:
            root_logger.error(f"Failed to configure file logging: {e}", exc_info=True)
    else:
        root_logger.info("Logging configured. Outputting to console only.")


class CheckpointManager:
    """ Handles saving and loading of model checkpoints. """
    
    @staticmethod
    def save(state: Dict[str, Any], file_path: Union[str, Path]) -> None:
        """ Save model state to file. """
        path = Path(file_path)
        path.parent.mkdir(parents=True, exist_ok=True)
        
        try:
            torch.save(state, path)
            logger.info(f"Saved checkpoint to {path}")
        except Exception as e:
            logger.error(f"Failed to save checkpoint: {e}", exc_info=True)

    @staticmethod
    def load(
        file_path: Union[str, Path],
        model: torch.nn.Module,
        optimizer: Optional[torch.optim.Optimizer] = None,
        map_location: Optional[Union[str, torch.device]] = None,
        strict: bool = True
    ) -> Dict[str, Any]:
        """ Load model state from file. """
        path = Path(file_path)
        if not path.exists():
            logger.error(f"Checkpoint not found: {path}")
            raise FileNotFoundError(f"Checkpoint not found: {path}")

        device = map_location or ('cuda' if torch.cuda.is_available() else 'cpu')
        logger.info(f"Loading checkpoint from {path} to {device}")

        checkpoint = torch.load(path, map_location=device)
        model.load_state_dict(checkpoint['model_state_dict'], strict=strict)
        logger.info("Model loaded")

        if optimizer and 'optimizer_state_dict' in checkpoint:
            try:
                optimizer.load_state_dict(checkpoint['optimizer_state_dict'])
                logger.info("Optimizer loaded")
            except Exception as e:
                logger.warning(f"Couldn't load optimizer: {e}")

        if metadata := {k: checkpoint.get(k) for k in ('epoch', 'loss') if k in checkpoint}:
            logger.info(f"Checkpoint metadata: {metadata}")

        return checkpoint
    

def save_df_to_csv(
    df: pd.DataFrame,
    output_name: str,
    output_dir: Union[str, Path] = './data/preprocessed',
    index: bool = False
) -> Path:
    """ Save DataFrame to CSV file, creating directory if needed. """
    if not isinstance(df, pd.DataFrame):
        raise TypeError("Input must be a pandas DataFrame")
    if not output_name:
        raise ValueError("Output name cannot be empty")

    # Ensure .csv extension
    output_name = output_name if output_name.endswith('.csv') else f"{output_name}.csv"
    
    file_path = Path(output_dir).resolve() / output_name
    file_path.parent.mkdir(parents=True, exist_ok=True)
    
    df.to_csv(file_path, index=index)
    logger.info(f"Saved DataFrame to {file_path}")
    return file_path


def minimal_clean_text(text: Any) -> str:
    """ Apply minimal cleaning to text. """
    if pd.isna(text):
        return ""
    if not isinstance(text, str):
        text = str(text)

    try:
        # 1. Unicode normalization (NFKC good for compatibility)
        text = unicodedata.normalize("NFKC", text)
        
        # 2. Normalize Arabic Alef forms [إأآا] -> ا
        text = re.sub(r"[إأآا]", "ا", text)
        
        # 3. Remove Tatweel (ـ) character used for justification
        text = re.sub(r"ـ", "", text)
        
        # 4. Remove Arabic diacritics/vowel marks (Tashkeel: ًٌٍَُِّْ)
        text = re.sub(r"[ًٌٍَُِّْ]", "", text)
        
        # 5. Collapse multiple whitespaces and strip leading/trailing whitespace
        text = re.sub(r'\s+', ' ', text).strip()

        # 6. Turn all characters to uppercase
        text = text.upper()
        
    except Exception as e:
        logger.warning(f"Error during text cleaning: {e}. Input: '{str(text)[:50]}...'")
        return ""  # Return empty string on error

    return text


def add_top_k_candidates_to_df(
    df: pd.DataFrame,
    ranking_results: List[List[Tuple[str, float, int]]],
    masterfile_df: pd.DataFrame,
    k: int = 3,
    include_cols: List[str] = ['sku', 'pred', 'sim', 'idx']
) -> pd.DataFrame:
    """ Add top-k ranked candidates as new columns to the input DataFrame. """
    # Input validation
    if len(ranking_results) != len(df):
        raise ValueError(f"Length mismatch: df has {len(df)} rows, ranking_results has {len(ranking_results)}")
    if k <= 0:
        return df.copy()
    if 'sku' in include_cols and 'sku' not in masterfile_df.columns:
        raise ValueError("Masterfile missing 'sku' column")

    # Create a mapping from index to SKU
    idx_to_sku = {}
    for i, (name_en, name_ar) in enumerate(zip(masterfile_df['product_name'], masterfile_df['product_name_ar'])):
        idx_to_sku[i] = masterfile_df.iloc[i]['sku']
        idx_to_sku[i + len(masterfile_df)] = masterfile_df.iloc[i]['sku']  # For Arabic names indices

    # Column generators
    col_gen = {
        'sku': lambda r: idx_to_sku.get(r[2]),
        'pred': lambda r: r[0],
        'sim': lambda r: round(r[1], 4),
        'idx': lambda r: r[2]
    }

    # Pre-allocate result DataFrame
    result_df = df.copy()
    
    # Process each rank and column type
    for rank in range(k):
        for col_type in include_cols:
            if col_type not in col_gen:
                continue
                
            col_name = f"{col_type}{rank+1}"
            result_df[col_name] = [
                col_gen[col_type](row[rank]) 
                if rank < len(row) and row[rank] 
                else None 
                for row in ranking_results
            ]

    return result_df


def segment_for_review(df, sim1_col='sim1', sim2_col='sim2', quantile=0.10, sim1_threshold=0.5):
    """
    Segment DataFrame into review and no-review parts based on similarity criteria.
    First filters by sim1_threshold, then applies quantile calculation on filtered data.
    
    Parameters:
    - df: DataFrame containing the data
    - sim1_col: Name of the first similarity column (default 'sim1')
    - sim2_col: Name of the second similarity column (default 'sim2')
    - quantile: Quantile value to use for threshold calculation (default 0.10)
    - sim1_threshold: Minimum similarity threshold for candidates to be reviewed (default 0.5)
    
    Returns:
    A tuple containing:
    - review_df: DataFrame of candidates that need review
    - no_review_df: DataFrame of candidates that don't need review
    """
    # First filter by similarity threshold
    above_threshold = df[df[sim1_col] > sim1_threshold]
    
    if len(above_threshold) > 0:
        # Calculate difference threshold based on quantile, only for data above sim1_threshold
        diff = above_threshold[sim1_col] - above_threshold[sim2_col]
        diff_threshold = diff.quantile(quantile)
        
        # Create final mask combining both conditions
        review_mask = (df[sim1_col] > sim1_threshold) & \
                     ((df[sim1_col] - df[sim2_col]) > diff_threshold)
    else:
        # If no data above threshold, use sim1_threshold only
        review_mask = df[sim1_col] > sim1_threshold
    
    # Segment the DataFrame
    no_review_df = df[review_mask].copy()
    review_df = df[~review_mask].copy()
    
    return review_df, no_review_df


def save_dataframe_to_xlsx(df: pd.DataFrame, file_path: str, sheet_name: str = "Sheet1") -> None:
    """ Save DataFrame to Excel file, handling both new and existing files. """
    logger = logging.getLogger(__name__)
    os.makedirs(os.path.dirname(file_path), exist_ok=True)
    
    try:
        # Try modern pandas approach first
        with pd.ExcelWriter(
            file_path,
            mode='a' if os.path.exists(file_path) else 'w',
            engine='openpyxl',
            if_sheet_exists='replace'
        ) as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            action = "appended to" if os.path.exists(file_path) else "saved to new"
            logger.info(f"DataFrame {action} {file_path} (sheet: '{sheet_name}')")
    
    except Exception as e:
        # Fallback for older pandas versions or special cases
        logger.warning(f"Modern save failed ({e}), using fallback method")
        if os.path.exists(file_path):
            book = load_workbook(file_path)
            if sheet_name in book.sheetnames:
                book.remove(book[sheet_name])
            book.save(file_path)
        
        df.to_excel(file_path, sheet_name=sheet_name, index=False, engine='openpyxl')
        logger.info(f"DataFrame saved to {file_path} using fallback method")